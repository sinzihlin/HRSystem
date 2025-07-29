import csv
import json
import logging
from datetime import datetime
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.views import LoginView, LogoutView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.views.generic import ListView, View, DetailView
from django.http import JsonResponse
from django.views.decorators.http import require_POST, require_GET
from django.utils import timezone
from django.db import connection
from django.core.exceptions import ImproperlyConfigured
from django.db.utils import DatabaseError, OperationalError
import openpyxl # 新增這一行

from .models import Department, Employee, Salary, Punch, Leave

# 設定日誌
logger = logging.getLogger(__name__)

def check_database_connection():
    """檢查資料庫連線是否正常"""
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT 1")
        return True
    except (DatabaseError, OperationalError) as e:
        logger.error(f"資料庫連線失敗: {e}")
        return False

class CustomLoginView(LoginView):
    template_name = 'employee/login.html'
    fields = '__all__'
    redirect_authenticated_user = True

    def form_valid(self, form):
        response = super().form_valid(form)
        try:
            employee = Employee.objects.get(email=self.request.user.email)
            self.request.session['employee_id'] = employee.id
        except Employee.DoesNotExist:
            self.request.session['employee_id'] = None
            messages.warning(self.request, "您的用戶帳號未綁定員工資料，部分功能可能受限。")
        return response

class CustomLogoutView(LogoutView):
    template_name = 'employee/logged_out.html'


@require_POST
def import_employees_api(request):
    if not request.user.is_authenticated or not request.user.is_staff:
        return JsonResponse({'status': 'error', 'message': '未授權。'}, status=403)

    # 檢查資料庫連線
    if not check_database_connection():
        return JsonResponse({
            'status': 'error', 
            'message': '資料庫暫時無法連線，請稍後再試。'
        }, status=503)

    uploaded_file = request.FILES.get('file') # 將 csv_file 改為 file 以支援多種檔案類型
    if not uploaded_file:
        return JsonResponse({'status': 'error', 'message': '請選擇一個檔案。'})

    file_extension = uploaded_file.name.split('.')[-1].lower()
    
    if file_extension not in ['csv', 'xlsx']:
        return JsonResponse({'status': 'error', 'message': '請上傳有效的 CSV 或 Excel 檔案 (.csv, .xlsx)。'})

    imported_count = 0
    updated_count = 0
    errors = []

    try:
        if file_extension == 'csv':
            decoded_file = uploaded_file.read().decode('utf-8').splitlines()
            reader = csv.DictReader(decoded_file)
            rows = list(reader)
        elif file_extension == 'xlsx':
            workbook = openpyxl.load_workbook(uploaded_file)
            sheet = workbook.active
            headers = [cell.value for cell in sheet[1]]
            rows = []
            for row_idx in range(2, sheet.max_row + 1):
                row_data = {}
                for col_idx, cell in enumerate(sheet[row_idx]):
                    row_data[headers[col_idx]] = cell.value
                rows.append(row_data)

        for i, row in enumerate(rows):
            try:
                # 檢查必要的欄位是否存在
                required_fields = ['employee_id', 'name', 'email', 'phone', 'department_name', 'hire_date']
                for field in required_fields:
                    if field not in row or row[field] is None:
                        raise KeyError(f"缺少欄位: {field}")

                department_name = row['department_name']
                department, created_dept = Department.objects.get_or_create(name=department_name)

                hire_date_str = str(row['hire_date']) # 確保是字串，以防 Excel 讀取為 datetime 物件
                # 處理 Excel 可能將日期讀取為 datetime 物件的情況
                if isinstance(row['hire_date'], datetime):
                    hire_date = row['hire_date'].date()
                else:
                    hire_date = datetime.strptime(hire_date_str, '%Y-%m-%d').date()

                employee, created_emp = Employee.objects.update_or_create(
                    employee_id=row['employee_id'], # 使用 employee_id 作為唯一識別碼
                    defaults={
                        'name': row['name'],
                        'email': row['email'],
                        'phone': row['phone'],
                        'department': department,
                        'hire_date': hire_date,
                    }
                )

                if created_emp:
                    imported_count += 1
                else:
                    updated_count += 1

            except KeyError as e:
                errors.append(f'第 {i+2} 行錯誤: 缺少欄位 {e}')
            except ValueError as e:
                errors.append(f'第 {i+2} 行錯誤: 日期格式不正確或資料無效 - {e}')
            except (DatabaseError, OperationalError) as e:
                logger.error(f'資料庫操作錯誤: {e}')
                errors.append(f'第 {i+2} 行錯誤: 資料庫操作失敗')
            except Exception as e:
                errors.append(f'第 {i+2} 行錯誤: {e}')

        if errors:
            return JsonResponse({'status': 'error', 'message': '匯入完成，但存在錯誤。', 'errors': errors})
        else:
            return JsonResponse({'status': 'success', 'message': f'成功匯入 {imported_count} 筆新員工資料，更新 {updated_count} 筆員工資料。'})
    
    except (DatabaseError, OperationalError) as e:
        logger.error(f'匯入員工資料時發生資料庫錯誤: {e}')
        return JsonResponse({
            'status': 'error', 
            'message': '資料庫操作失敗，請檢查資料庫連線或聯絡系統管理員。'
        }, status=503)
    except Exception as e:
        logger.error(f'匯入員工資料時發生未預期錯誤: {e}')
        return JsonResponse({
            'status': 'error', 
            'message': f'匯入失敗: {str(e)}'
        }, status=500)

@require_POST
def import_punches_excel_api(request):
    if not request.user.is_authenticated or not request.user.is_staff:
        return JsonResponse({'status': 'error', 'message': '未授權。'}, status=403)

    if not check_database_connection():
        return JsonResponse({
            'status': 'error',
            'message': '資料庫暫時無法連線，請稍後再試。'
        }, status=503)

    uploaded_file = request.FILES.get('file')
    if not uploaded_file:
        return JsonResponse({'status': 'error', 'message': '請選擇一個檔案。'}) 

    file_extension = uploaded_file.name.split('.')[-1].lower()
    if file_extension != 'xlsx':
        return JsonResponse({'status': 'error', 'message': '請上傳有效的 Excel 檔案 (.xlsx)。'}) 

    imported_count = 0
    errors = []

    try:
        workbook = openpyxl.load_workbook(uploaded_file)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]

        # Expected headers for punch data
        expected_headers = ['employee_id', 'punch_type', 'punch_time'] # 將 employee_email 改為 employee_id
        if not all(h in headers for h in expected_headers):
            return JsonResponse({'status': 'error', 'message': f'Excel 檔案缺少必要的欄位。需要: {", ".join(expected_headers)}'}, status=400)

        for row_idx in range(2, sheet.max_row + 1):
            row_data = {}
            for col_idx, cell in enumerate(sheet[row_idx]):
                if col_idx < len(headers):
                    row_data[headers[col_idx]] = cell.value

            try:
                employee_id = row_data.get('employee_id') # 從 employee_email 改為 employee_id
                punch_type = row_data.get('punch_type')
                punch_time_value = row_data.get('punch_time')

                if not employee_id or not punch_type or not punch_time_value:
                    raise ValueError("缺少必要的打卡資訊 (員工編號、打卡類型或打卡時間)。")

                employee = Employee.objects.get(employee_id=employee_id) # 從 email 改為 employee_id

                # Handle various date/time formats from Excel
                if isinstance(punch_time_value, datetime):
                    punch_time = timezone.make_aware(punch_time_value)
                elif isinstance(punch_time_value, str):
                    try:
                        punch_time = timezone.make_aware(datetime.strptime(punch_time_value, '%Y-%m-%d %H:%M:%S'))
                    except ValueError:
                        punch_time = timezone.make_aware(datetime.strptime(punch_time_value, '%Y-%m-%d %H:%M'))
                else:
                    raise ValueError("無法解析打卡時間格式。")

                Punch.objects.create(
                    employee=employee,
                    punch_time=punch_time,
                    punch_type=punch_type
                )
                imported_count += 1

            except Employee.DoesNotExist:
                errors.append(f'第 {row_idx} 行錯誤: 員工編號 {employee_id} 不存在。') # 錯誤訊息調整
            except ValueError as e:
                errors.append(f'第 {row_idx} 行錯誤: 資料格式不正確或無效 - {e}')
            except (DatabaseError, OperationalError) as e:
                logger.error(f'資料庫操作錯誤: {e}')
                errors.append(f'第 {row_idx} 行錯誤: 資料庫操作失敗')
            except Exception as e:
                errors.append(f'第 {row_idx} 行錯誤: {e}')

        if errors:
            return JsonResponse({'status': 'error', 'message': '匯入完成，但存在錯誤。', 'errors': errors})
        else:
            return JsonResponse({'status': 'success', 'message': f'成功匯入 {imported_count} 筆打卡記錄。'})

    except Exception as e:
        logger.error(f'匯入打卡記錄時發生未預期錯誤: {e}')
        return JsonResponse({
            'status': 'error',
            'message': f'匯入失敗: {str(e)}'
        }, status=500)

@require_POST
def punch_api(request):
    logger.debug("收到打卡請求: %s", request.body.decode('utf-8'))
    if not request.user.is_authenticated or not request.user.is_staff:
        logger.warning("未授權的打卡請求，用戶: %s", request.user)
        return JsonResponse({'status': 'error', 'message': '未授權。'}, status=403)

    try:
        data = json.loads(request.body)
        employee_id = data.get('employee_id')
        punch_type = data.get('punch_type')
        punch_time_str = data.get('punch_time')

        if not employee_id or not punch_type:
            return JsonResponse({'status': 'error', 'message': '缺少員工 ID 或打卡類型。'})

        employee = Employee.objects.get(id=employee_id)

        if punch_time_str:
            # 假設傳入的日期時間格式為 YYYY-MM-DD HH:MM:SS
            punch_time = datetime.strptime(punch_time_str, '%Y-%m-%d %H:%M:%S')
            # 將 datetime 物件轉換為帶有時區資訊的 datetime 物件
            punch_time = timezone.make_aware(punch_time)
        else:
            punch_time = timezone.now()

        Punch.objects.create(
            employee=employee,
            punch_time=punch_time,
            punch_type=punch_type
        )
        return JsonResponse({'status': 'success', 'message': f'{employee.name} 已成功打卡 ({punch_type})。'})

    except Employee.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': '員工不存在。'})
    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': '無效的 JSON 格式。'})
    except ValueError as e:
        return JsonResponse({'status': 'error', 'message': f'日期時間格式錯誤: {e}'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': f'打卡失敗: {e}'})

@require_POST
def leave_api(request):
    if not request.user.is_authenticated or not request.user.is_staff:
        return JsonResponse({'status': 'error', 'message': '未授權。'}, status=403)

    try:
        data = json.loads(request.body)
        employee_id = data.get('employee_id')
        leave_type = data.get('leave_type')
        start_date_str = data.get('start_date')
        end_date_str = data.get('end_date')
        reason = data.get('reason', '')

        if not employee_id or not leave_type or not start_date_str or not end_date_str:
            return JsonResponse({'status': 'error', 'message': '缺少必要的請假資訊。'})

        employee = Employee.objects.get(id=employee_id)
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

        if start_date > end_date:
            return JsonResponse({'status': 'error', 'message': '開始日期不能晚於結束日期。'})

        Leave.objects.create(
            employee=employee,
            leave_type=leave_type,
            start_date=start_date,
            end_date=end_date,
            reason=reason,
            status='PENDING' # 預設為待審核
        )
        return JsonResponse({'status': 'success', 'message': f'{employee.name} 的 {leave_type} 請假申請已提交。'})

    except Employee.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': '員工不存在。'})
    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': '無效的 JSON 格式。'})
    except ValueError as e:
        return JsonResponse({'status': 'error', 'message': f'日期格式錯誤: {e}'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': f'提交請假失敗: {e}'})

@require_GET
def get_pending_leaves_api(request):
    if not request.user.is_authenticated or not request.user.is_staff:
        return JsonResponse({'status': 'error', 'message': '未授權。'}, status=403)

    pending_leaves = Leave.objects.filter(status='PENDING').order_by('-applied_date')
    data = []
    for leave in pending_leaves:
        data.append({
            'id': leave.id,
            'employee_name': leave.employee.name,
            'leave_type': leave.get_leave_type_display(),
            'start_date': leave.start_date.strftime('%Y-%m-%d'),
            'end_date': leave.end_date.strftime('%Y-%m-%d'),
            'reason': leave.reason,
            'applied_date': leave.applied_date.strftime('%Y-%m-%d'),
        })
    return JsonResponse({'status': 'success', 'leaves': data})

@require_POST
def update_leave_status_api(request):
    if not request.user.is_authenticated or not request.user.is_staff:
        return JsonResponse({'status': 'error', 'message': '未授權。'}, status=403)

    try:
        data = json.loads(request.body)
        leave_id = data.get('leave_id')
        new_status = data.get('status')

        if not leave_id or not new_status or new_status not in ['APPROVED', 'REJECTED']:
            return JsonResponse({'status': 'error', 'message': '缺少請假 ID 或無效的狀態。'})

        leave = Leave.objects.get(id=leave_id)
        leave.status = new_status
        leave.save()

        return JsonResponse({'status': 'success', 'message': f'請假申請已更新為 {leave.get_status_display()}。'})

    except Leave.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': '請假申請不存在。'})
    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': '無效的 JSON 格式。'})
    except ValueError as e:
        return JsonResponse({'status': 'error', 'message': f'日期時間格式錯誤: {e}'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': f'更新請假狀態失敗: {e}'})


from django.shortcuts import get_object_or_404

def employee_punches_view(request, employee_id):
    employee = get_object_or_404(Employee, pk=employee_id)
    punches = Punch.objects.filter(employee=employee).order_by('-punch_time')
    leaves = Leave.objects.filter(employee=employee).order_by('-applied_date')
    employees = Employee.objects.all().order_by('name') # 獲取所有員工
    return render(request, 'employee/employee_punches.html', {'employee': employee, 'punches': punches, 'leaves': leaves, 'employees': employees})

@require_GET
def health_check(request):
    """系統健康檢查端點"""
    try:
        # 檢查資料庫連線
        db_status = check_database_connection()
        
        # 檢查基本資料表是否存在
        tables_exist = True
        try:
            Employee.objects.count()
            Department.objects.count()
        except (DatabaseError, OperationalError):
            tables_exist = False
        
        status = 'healthy' if db_status and tables_exist else 'unhealthy'
        
        response_data = {
            'status': status,
            'database_connection': db_status,
            'tables_exist': tables_exist,
            'timestamp': timezone.now().isoformat()
        }
        
        if status == 'healthy':
            response_data['employee_count'] = Employee.objects.count()
            response_data['department_count'] = Department.objects.count()
        
        return JsonResponse(response_data)
    
    except Exception as e:
        logger.error(f'健康檢查時發生錯誤: {e}')
        return JsonResponse({
            'status': 'unhealthy',
            'error': str(e),
            'timestamp': timezone.now().isoformat()
        }, status=500)

def settings_view(request):
    """渲染設定頁面"""
    return render(request, 'employee/settings.html')